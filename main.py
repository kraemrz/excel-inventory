#!/usr/bin/env python3
import openpyxl
import argparse
import os


# Script to compare parts physical in stock and whats in database,
# take a database file as input and physical file as commpare.
# databse and physical file must have partnr in colum 1
# and physical must have have quantity in column 2.
# this script must be in same folder as .xlsx files
def main():
    parser = argparse.ArgumentParser(
        description="Takes database file as input and physical \
        counted as compare file and puts out a updated.xlsx")
    parser.add_argument("-i", "--input", type=str, metavar="",
                        required=True, help="Name of file from db")
    parser.add_argument("-p", "--physical", type=str, metavar="",
                        required=True, help="Name of physical file")
    args = parser.parse_args()

    # get directory of script
    cwd = os.getcwd() + "/"
    input_file = (cwd + args.infile)

    # creating a dict of physical parts
    parts = parts_in_car(cwd + args.outfile)

    # loading database file
    input_workbook = openpyxl.load_workbook(input_file)
    sheet = input_workbook.active
    sheet2 = input_workbook.create_sheet("Not in database")

    # looping through db file and if art number in dict update column E
    # and removing key:value from dict
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value in parts.keys():
            sheet['E' + str(i)].value = parts[sheet.cell(row=i,
                                                         column=1).value]
            del parts[sheet.cell(row=i, column=1).value]
        else:
            continue

    # creating sheet 2 for not found parts
    # sheet2['A1'] = "artnr"
    # sheet2['B2'] = "antal"
    for row, (artnr, qty) in enumerate(parts.items(), start=2):
        sheet2[f'A{row}'] = artnr
        sheet2[f'A{row}'] = qty

    input_workbook.save('updated.xlsx')


# function to create dict of physical parts
def parts_in_car(inv_file):
    inv_workbook = openpyxl.load_workbook(inv_file)
    inv_sheet = inv_workbook.active

    inventoried_parts = {}
    for i in range(1, inv_sheet.max_row + 1):
        inventoried_parts.update({
            inv_sheet.cell(row=i, column=1).value:
            inv_sheet.cell(row=i, column=2).value})

    return inventoried_parts


if __name__ == '__main__':
    main()
